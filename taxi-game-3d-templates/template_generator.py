from openpyxl import load_workbook

def generate_car(file_path, sheet_name):
    book = load_workbook(file_path, data_only=True)
    sheet = book[sheet_name]
    skip = 0
    temp_group = []
    for row in sheet.rows:
        # 세번째 줄부터 데이터 저장
        if skip < 2:
            skip += 1
            continue      
        # ID컬럼이 비어 있으면 추가하지 않음
        if row[0].value == None:
            continue
        new_temp = {
            'Id': row[0].value,
            'Name': {
                'Table': row[1].value,
                'Key': row[2].value
            },
            'Icon': row[3].value,
            'Prefab': row[4].value,
            'Cost': int(row[5].value)
        }
        temp_group.append(new_temp)
    return temp_group

def generate_customer(file_path, sheet_name):
    book = load_workbook(file_path, data_only=True)
    sheet = book[sheet_name]
    skip = 0
    temp_group = []
    for row in sheet.rows:
        # 두번째 줄부터 데이터 저장
        if skip < 1:
            skip += 1
            continue   
        # ID컬럼이 비어 있으면 추가하지 않음
        if row[0].value == None:
            continue
        new_temp = {
            'Id': row[0].value,
            'Icon': row[1].value,
            'Prefab': row[2].value,
        }
        temp_group.append(new_temp)
    return temp_group

def generate_stage(file_path, sheet_name):
    book = load_workbook(file_path, data_only=True)
    sheet = book[sheet_name]
    skip = 0
    temp_group = []
    for row in sheet.rows:
        # 두번째 줄부터 데이터 저장
        if skip < 1:
            skip += 1
            continue   
        # ID컬럼이 비어 있으면 추가하지 않음
        if row[0].value == None:
            continue
        new_temp = {
            'Id': row[0].value,
            'Scene': row[1].value,
            'Distance': float(row[2].value),
            'FareRate': float(row[3].value)
        }
        temp_group.append(new_temp)
    return temp_group

def generate_talk(file_path, sheet_name):
    book = load_workbook(file_path, data_only=True)
    sheet = book[sheet_name]
    skip = 0
    temp_group = []
    for row in sheet.rows:
        # 세번째 줄부터 데이터 저장
        if skip < 2:
            skip += 1
            continue   
        # ID컬럼이 비어 있으면 추가하지 않음
        if row[0].value == None:
            continue
        new_temp = {
            'Type': int(row[0].value),
            'Content': {
                'Table': row[1].value,
                'Key': row[2].value
            }
        }
        temp_group.append(new_temp)
    return temp_group
